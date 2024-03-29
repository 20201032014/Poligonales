import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import math
import shutil

usersDB = [
  {
    "user": 'Pepito',
    "password": '123'
  },
  {
    "user": 'Juan',
    "password": '158'
  },
  {
    "user": 'Pepe',
    "password": '948'
  },
]

procesoTerminado = False
while (True):
  if (procesoTerminado):
    break
  userInput = input('Ingrese Usuario: ')
  passwordInput = input('Ingrese Contraseña: ')

  userExist = False
  for userDB in usersDB:
    if (userDB['user'] == userInput):
      userExist = True
      if (userDB['password'] == passwordInput):
        print('-- Logeado --')
        while (True):
          print("""
  Correccon de poligonales:
      1. Brujula
      2. Transito
      3. Crandall
  Nivelaciones
      4. Geometrica
      5. Trigonometrica
          """)
          metodoInput = input('Ingrese Metodo: ')
          if (metodoInput == '1' or metodoInput == '2' or metodoInput == '3' or metodoInput == '4' or metodoInput == '5' or metodoInput == 'Brujula' or metodoInput == 'Transito' or metodoInput == 'Crandall' or metodoInput == 'Geometrica' or metodoInput == 'Trigonometrica'):
            break
          else:
            print('Ingrese un Metodo Correcto')

        # Init
        if (metodoInput != 'Geometrica' and metodoInput != '4' and metodoInput != 'Trigonometrica' and metodoInput != '5'):
          norteInput = int(input('Ingrese Norte: '))
          esteInput = int(input('Ingrese Este: '))
          shutil.copy('./Excel.xlsx', './ExcelGenerado.xlsx')
          wb = openpyxl.load_workbook('./ExcelGenerado.xlsx')
        if (metodoInput == 'Geometrica' or metodoInput == '4'):
          shutil.copy('./NC.xlsx', './NCGenerado.xlsx')
          wb = openpyxl.load_workbook('./NCGenerado.xlsx')
        if (metodoInput == 'Trigonometrica' or metodoInput == '5'):
          shutil.copy('./NV.xlsx', './NVGenerado.xlsx')
          wb = openpyxl.load_workbook('./NVGenerado.xlsx')

        if (metodoInput == 'Brujula' or metodoInput == '1'):
          print('Brujula')
          sheet = wb['Brujula']
          rows = sheet.max_row
          columns = sheet.max_column

          rowsExcel = []

          # FOR's para guardar los datos del Excel en un array
          for row in range(2, rows):
            rowArr = []
            for column in range(columns):
              valueCell = sheet.cell(row=row+1,column=column+1).value
              rowArr.append(valueCell)

            if(rowArr[0] == None and rowArr[1] != None):
              DH = rowArr[6]
              # Calcular Azimut
              grados = rowArr[2]
              minutos = rowArr[3]
              segundos = rowArr[4]
              azimut = math.radians(grados+(minutos/60)+(segundos/3600))
              # Almacenar Azimut en rowArr, para despues guardar en Excel
              rowArr[5] = azimut
              # Calcular Y
              Y = math.cos(azimut) * DH
              X = math.sin(azimut) * DH
              rowArr[7] = Y
              rowArr[8] = X

            if (rowArr[0] == None and rowArr[1] != None):
              rowsExcel.append(rowArr)

          # Calcular Sumatorias DH Y X
          sumDH = 0
          sumY = 0
          sumX = 0
          for row in rowsExcel:
            if (row[0] == None):
              sumDH += row[6]
              sumY += row[7]
              sumX += row[8]

          # Calcular CY - CX - YC - XC
          for row in range(len(rowsExcel)):
            if (rowsExcel[row][0] == None):
              DH = rowsExcel[row][6]
              CY = sumY / sumDH * DH
              CX = sumX / sumDH * DH

              rowsExcel[row][9] = CY
              rowsExcel[row][10] = CX

              YC = rowsExcel[row][7] - CY
              XC = rowsExcel[row][8] - CX

              rowsExcel[row][11] = YC
              rowsExcel[row][12] = XC

          for row in range(len(rowsExcel)):
            # print(row, '***>', rowsExcel[row])
            if (row == 0):
              rowsExcel[row][13] = rowsExcel[row][11] + norteInput
              rowsExcel[row][14] = rowsExcel[row][12] + esteInput
            else:
              if (rowsExcel[row][1] != None):
                rowsExcel[row][13] = rowsExcel[row][11] + rowsExcel[row - 1][13]
                rowsExcel[row][14] = rowsExcel[row][12] + rowsExcel[row - 1][14]

          # for row in range(len(rowsExcel)):
          #   print(rowsExcel[row])

          count4 = 4
          for row in range(len(rowsExcel)):
            sheet[f'N{count4}'] = rowsExcel[row][13]
            sheet[f'O{count4}'] = rowsExcel[row][14]
            count4+=2
          
          sheet['N3'] = norteInput
          sheet['O3'] = esteInput

          wb.save('ExcelGenerado.xlsx')
          print('Datos generados en ExcelGenerado.xlsx')

          xxx = [norteInput]
          yyy = [esteInput]
          for rowG in rowsExcel:
            if (rowG[13] != None):
              xxx.append(rowG[13])
              yyy.append(rowG[14])

          # * Grafica
          fig, ax = plt.subplots()
          ax.set_title(f'Grafica Correccion de Poligonal')
          ax.plot(xxx, yyy)
          plt.show()

        if (metodoInput == 'Transito' or metodoInput == '2'):
          print('Transito')
          sheet = wb['Transito']
          rows = sheet.max_row
          columns = sheet.max_column

          count = 1
          rowsExcel = []

          # FOR's para guardar los datos del Excel en un array
          for row in range(2, rows):
            rowArr = []
            for column in range(columns):
              valueCell = sheet.cell(row=row+1,column=column+1).value
              rowArr.append(valueCell)

            if(rowArr[0] == None and rowArr[1] != None):
              DH = rowArr[6]
              # Calcular Azimut
              grados = rowArr[2]
              minutos = rowArr[3]
              segundos = rowArr[4]
              azimut = math.radians(grados+(minutos/60)+(segundos/3600))
              # Almacenar Azimut en rowArr, para despues guardar en Excel
              rowArr[5] = azimut
              # Calcular Y
              Y = math.cos(azimut) * DH
              X = math.sin(azimut) * DH
              rowArr[7] = Y
              rowArr[9] = X

            if (rowArr[0] == None and rowArr[1] != None):
              rowsExcel.append(rowArr)

          # Calcular Sumatorias DH Y X
          sumDH = 0
          sumY = 0
          sumX = 0
          for row in rowsExcel:
            if (row[0] == None):
              sumDH += row[6]
              sumY += row[7]
              sumX += row[9]

          # Calcular VY VX
          for row in range(len(rowsExcel)):
            if (rowsExcel[row][0] == None):
              rowsExcel[row][8] = abs(rowsExcel[row][7])
              rowsExcel[row][10] = abs(rowsExcel[row][9])

          # Sumatorias VY VX
          sumVY = 0
          sumVX = 0
          for row in rowsExcel:
            if (row[0] == None):
              sumVY += row[8]
              sumVX += row[10]

          # Calcular CY - CX - YC - XC
          for row in range(len(rowsExcel)):
            if (rowsExcel[row][0] == None):
              DH = rowsExcel[row][6]
              CY = (sumY / sumVY) * rowsExcel[row][7]
              CX = (sumX / sumVX) * rowsExcel[row][9]

              rowsExcel[row][11] = CY
              rowsExcel[row][12] = CX

              if (rowsExcel[row][7] < 0):
                YC = (rowsExcel[row][8] - CY) * -1
              else:
                YC = rowsExcel[row][8] - CY

              if (rowsExcel[row][9] < 0):
                XC = (rowsExcel[row][10] - CX) * -1
              else:
                XC = rowsExcel[row][10] - CX

              rowsExcel[row][13] = YC
              rowsExcel[row][14] = XC

          for row in range(len(rowsExcel)):
            if (row == 0):
              rowsExcel[row][15] = rowsExcel[row][13] + norteInput
              rowsExcel[row][16] = rowsExcel[row][14] + esteInput
            else:
              if (rowsExcel[row][1] != None):
                rowsExcel[row][15] = rowsExcel[row][13] + rowsExcel[row - 1][15]
                rowsExcel[row][16] = rowsExcel[row][14] + rowsExcel[row - 1][16]

          # for row in range(len(rowsExcel)):
          #   print(rowsExcel[row])

          count4 = 4
          for row in range(len(rowsExcel)):
            sheet[f'P{count4}'] = rowsExcel[row][15]
            sheet[f'Q{count4}'] = rowsExcel[row][16]
            count4+=2

          sheet['P3'] = norteInput
          sheet['Q3'] = norteInput

          wb.save('ExcelGenerado.xlsx')
          print('Datos generados en ExcelGenerado.xlsx')

          xxx = [norteInput]
          yyy = [esteInput]
          for rowG in rowsExcel:
            if (rowG[13] != None):
              xxx.append(rowG[15])
              yyy.append(rowG[16])

          # * Grafica
          fig, ax = plt.subplots()
          ax.set_title(f'Grafica Correccion de Poligonal')
          ax.plot(xxx, yyy)
          plt.show()

        if (metodoInput == 'Crandall' or metodoInput == '3'):
          print('Crandall')
          sheet = wb['Crandall']
          rows = sheet.max_row
          columns = sheet.max_column

          count = 1
          rowsExcel = []

          # FOR's para guardar los datos del Excel en un array
          for row in range(2, rows):
            rowArr = []
            for column in range(columns):
              valueCell = sheet.cell(row=row+1,column=column+1).value
              rowArr.append(valueCell)

            if(rowArr[0] == None and rowArr[1] != None):
              DH = rowArr[6]
              # Calcular Azimut
              grados = rowArr[2]
              minutos = rowArr[3]
              segundos = rowArr[4]
              azimut = math.radians(grados+(minutos/60)+(segundos/3600))
              # Almacenar Azimut en rowArr, para despues guardar en Excel
              rowArr[5] = azimut
              # Calcular Y
              Y = math.cos(azimut) * DH
              X = math.sin(azimut) * DH
              rowArr[7] = Y
              rowArr[8] = X

            if (rowArr[0] == None and rowArr[1] != None):
              rowsExcel.append(rowArr)

          # Calcular Sumatorias DH Y X
          sumDH = 0
          sumY = 0
          sumX = 0
          for row in rowsExcel:
            if (row[0] == None):
              sumDH += row[6]
              sumY += row[7]
              sumX += row[8]

          # Calcular L^2 D^2 LD LD^2
          for row in rowsExcel:
            if (row[0] == None):
              L2 = ((row[7] * row[7]) / row[6]) / 100
              D2 = ((row[8] * row[8]) / row[6]) / 100
              LD = ((row[7] + row[8]) / row[6]) / 100
              LD2 = LD * LD

              row[9] = L2
              row[10] = D2
              row[11] = LD
              row[12] = LD2

          # Calcular Sumatorias L^2 D^2 LD LD^2
          sumL2 = 0
          sumD2 = 0
          sumLD = 0
          sumLD2 = 0
          for row in rowsExcel:
            if (row[0] == None):
              sumL2 += row[9]
              sumD2 += row[10]
              sumLD += row[11]
              sumLD2 += row[12]

          # Calcular A B
          A = ((sumX * sumLD) - (sumY * sumD2)) / ((sumL2 * sumD2) - sumLD2)
          B = ((sumY * sumLD) - (sumX * sumL2)) / ((sumL2 * sumD2) - sumLD2)

          # Calcular CX CY YCORREGIDO XCORREGIDO
          for row in rowsExcel:
            if (row[0] == None):
              CY = (row[9] * A) + (row[11] * B)
              CX = (row[10] * A) + (row[11] * B)
              row[13] = CY
              row[14] = CX

              YCORREGIDO = row[7] + CY
              XCORREGIDO = row[8] + CX
              row[15] = YCORREGIDO
              row[16] = XCORREGIDO

          # for row in range(len(rowsExcel)):
          #   print(rowsExcel[row][7], rowsExcel[row][8])

          for row in range(len(rowsExcel)):
            if (row == 0):
              rowsExcel[row][17] = rowsExcel[row][15] + norteInput
              rowsExcel[row][18] = rowsExcel[row][16] + esteInput
            else:
              if (rowsExcel[row][1] != None):
                rowsExcel[row][17] = rowsExcel[row][15] + rowsExcel[row - 1][17]
                rowsExcel[row][18] = rowsExcel[row][16] + rowsExcel[row - 1][18]
          
          count4 = 4
          for row in range(len(rowsExcel)):
            # print(rowsExcel[row])
            sheet[f'R{count4}'] = rowsExcel[row][17]
            sheet[f'S{count4}'] = rowsExcel[row][18]
            count4+=2
          
          sheet['R3'] = norteInput
          sheet['S3'] = norteInput

          wb.save('ExcelGenerado.xlsx')
          print('Datos generados en ExcelGenerado.xlsx')

          xxx = [norteInput]
          yyy = [esteInput]
          for rowG in rowsExcel:
            if (rowG[13] != None):
              xxx.append(rowG[17])
              yyy.append(rowG[18])

          # * Grafica
          fig, ax = plt.subplots()
          ax.set_title(f'Grafica Correccion de Poligonal')
          ax.plot(xxx, yyy)
          plt.show()

        if (metodoInput == 'Geometrica' or metodoInput == '4'):
          print('Geometrica')
          sheet = wb['Hoja1']
          rows = sheet.max_row
          columns = sheet.max_column

          count = 1
          rowsExcel = []

          # FOR's para guardar los datos del Excel en un array
          for row in range(1, rows-1):
            rowArr = []
            for column in range(columns):
              valueCell = sheet.cell(row=row+1,column=column+1).value
              rowArr.append(valueCell)

            rowsExcel.append(rowArr)
            count+=1

          # Ver Inicio y Fin de Puntos
          count2 = 0
          initPuntos1 = []
          finPuntos1 = []
          for row in rowsExcel:
            # Punto Nuevo
            if (row[0] != None and row[1] != None):
              initPuntos1.append(count2)
            if (row[0] == None and row[1] == None):
              for ip in range(1, len(initPuntos1)):
                finPuntos1.append(initPuntos1[ip]-1)
              finPuntos1.append(count2-1)
              break
            count2+=1

          # ! Ver Inicio y Fin de Puntos 2
          count3 = 0
          initPuntos2 = []
          finPuntos2 = []
          for row in rowsExcel:
            # Punto Nuevo
            if (row[11] != None and row[12] != None):
              initPuntos2.append(count3)
            if (row[11] == None and row[12] == None):
              for ip in range(1, len(initPuntos2)):
                finPuntos2.append(initPuntos2[ip]-1)
              finPuntos2.append(count3-1)
              break
            count3+=1

          # sum
          sumDistancias = 0
          sumDist = 0
          sumVistaP = 0
          sumVistaN = 0
          for i in range(len(initPuntos1)):
            sumVistaP = rowsExcel[initPuntos1[i]][4] + sumVistaP
            sumVistaN = rowsExcel[finPuntos1[i]][6] + sumVistaN
            # print(initPuntos1[i], finPuntos1[i])
            for ii in range(initPuntos1[i]+1, finPuntos1[i]+1):
              if (i == 0 and ii == 1):
                sumDistancias = rowsExcel[ii][2] + 0
                rowsExcel[ii][3] = sumDistancias
                sheet[f'D{ii+2}'] = sumDistancias
              else:
                sumDistancias = rowsExcel[ii][2] + sumDistancias
                rowsExcel[ii][3] = sumDistancias
                sheet[f'D{ii+2}'] = sumDistancias
              # print('-->:', rowsExcel[ii][2], rowsExcel[ii][3])
            sumDist = sumDistancias
          sheet[f'C{finPuntos1[-1]+3}'] = sumDist
          sheet[f'E{finPuntos1[-1]+3}'] = sumVistaP
          sheet[f'G{finPuntos1[-1]+3}'] = sumVistaN

          # Calcular primera Altura
          AI = rowsExcel[0][4] + rowsExcel[0][7]
          # print(AI)
          sheet['F2'] = AI

          # Cotas
          for i in range(len(initPuntos1)):
            if (i != 0):  
              AII = rowsExcel[initPuntos1[i]][4] + rowsExcel[initPuntos1[i]-1][7]
              # print(AII)
              sheet[f'F{initPuntos1[i]+2}'] = AII
            # print(initPuntos1[i], finPuntos1[i])
            for ii in range(initPuntos1[i]+1, finPuntos1[i]+1):
              if (i == 0):  
                rowsExcel[ii][7] = AI - rowsExcel[ii][6]
                sheet[f'H{ii+2}'] = rowsExcel[ii][7]
              else:
                rowsExcel[ii][7] = AII - rowsExcel[ii][6]
                sheet[f'H{ii+2}'] = rowsExcel[ii][7]

          # ! sum
          sumDistancias2 = 0
          sumDist2 = 0
          sumVistaP2 = 0
          sumVistaN2 = 0
          for i in range(len(initPuntos2)):
            sumVistaP2 = rowsExcel[initPuntos2[i]][15] + sumVistaP2
            sumVistaN2 = rowsExcel[finPuntos2[i]][17] + sumVistaN2
            # print(initPuntos2[i], finPuntos2[i])
            for ii in range(initPuntos2[i]+1, finPuntos2[i]+1):
              if (i == 0 and ii == 1):
                sumDistancias2 = rowsExcel[ii][13] + 0
                rowsExcel[ii][14] = sumDistancias2
                sheet[f'O{ii+2}'] = sumDistancias2
              else:
                sumDistancias2 = rowsExcel[ii][13] + sumDistancias2
                rowsExcel[ii][14] = sumDistancias2
                sheet[f'O{ii+2}'] = sumDistancias2
              # print('-->:', rowsExcel[ii][2], rowsExcel[ii][3])
            sumDist2 = sumDistancias
          sheet[f'N{finPuntos2[-1]+3}'] = sumDist2
          sheet[f'P{finPuntos2[-1]+3}'] = sumVistaP2
          sheet[f'R{finPuntos2[-1]+3}'] = sumVistaN2

          # ! Calcular primera Altura
          AI2 = rowsExcel[0][15] + rowsExcel[0][18]
          # print(AI2)
          sheet['Q2'] = AI2

          # ! Cotas
          for i in range(len(initPuntos2)):
            if (i != 0):  
              AII2 = rowsExcel[initPuntos2[i]][15] + rowsExcel[initPuntos2[i]-1][18]
              # print(AI2)
              sheet[f'Q{initPuntos2[i]+2}'] = AII2
            # print(initPuntos2[i], finPuntos2[i])
            for ii in range(initPuntos2[i]+1, finPuntos2[i]+1):
              if (i == 0):
                # print(rowsExcel[ii])
                rowsExcel[ii][18] = AI2 - rowsExcel[ii][17]
                sheet[f'S{ii+2}'] = rowsExcel[ii][18]
              else:
                rowsExcel[ii][18] = AII2 - rowsExcel[ii][17]
                sheet[f'S{ii+2}'] = rowsExcel[ii][18]
          
          # ? VistaNTotal VistaPTotal Error de Cierre
          VistaPTotal = sumVistaP + sumVistaP2
          sheet['AD2'] = VistaPTotal
          VistaNTotal = sumVistaN + sumVistaN2
          sheet['AD4'] = VistaNTotal
          ErrorCierre1 = VistaPTotal - VistaNTotal
          sheet['AJ8'] = ErrorCierre1
          ErrorCierre2 = rowsExcel[finPuntos2[-1]][18] - rowsExcel[0][7]
          sheet['AD8'] = ErrorCierre2


          # EC Cota Compensada
          for i in range(len(initPuntos1)):
            for ii in range(initPuntos1[i]+1, finPuntos1[i]+1):
              EC1 = (rowsExcel[ii][3] * ErrorCierre1) / sumDist
              CComp = EC1 + rowsExcel[ii][7]
              rowsExcel[ii][8] = EC1
              rowsExcel[ii][9] = CComp
              sheet[f'I{ii+2}'] = EC1
              sheet[f'J{ii+2}'] = CComp

          # ! EC Cota Compensada
          for i in range(len(initPuntos2)):
            for ii in range(initPuntos2[i]+1, finPuntos2[i]+1):
              EC2 = (rowsExcel[ii][14] * ErrorCierre1) / sumDist2
              CComp2 = EC2 + rowsExcel[ii][18]
              rowsExcel[ii][19] = EC2
              rowsExcel[ii][20] = CComp2
              sheet[f'T{ii+2}'] = EC2
              sheet[f'U{ii+2}'] = CComp2

          sheet[f'AC13'] = sumDist
          sheet[f'AC15'] = sumDist2
          sheet[f'AC17'] = sumDist + sumDist2

          wb.save('NCGenerado.xlsx')
          print('Datos generados en NCGenerado.xlsx')

        if (metodoInput == 'Trigonometrica' or metodoInput == '5'):
          print('Trigonometrica')
          sheet = wb['Hoja1']
          rows = sheet.max_row
          columns = sheet.max_column

          count = 1
          rowsExcel = []

          # FOR's para guardar los datos del Excel en un array
          for row in range(1, rows-1):
            rowArr = []
            for column in range(columns):
              valueCell = sheet.cell(row=row+1,column=column+1).value
              rowArr.append(valueCell)

            rowsExcel.append(rowArr)
            count+=1

          # Angulo V
          for i in range(len(rowsExcel)-1):
            print(rowsExcel[i][2], rowsExcel[i][3], rowsExcel[i][4])
            AV = rowsExcel[i][2] + (rowsExcel[i][3]/60) + (rowsExcel[i][4]/3600)
            rowsExcel[i][5] = AV
            sheet[f'F{i+2}'] = AV

            # * DH
            print(rowsExcel[i][6], rowsExcel[i][8])
            if (AV >= 90):
              DH = 100*(rowsExcel[i][6]-rowsExcel[i][8])*(math.cos(AV) * math.cos(AV))
            else:
              DH = 100*(rowsExcel[i][6]-rowsExcel[i][8])*(math.cos(90-AV) * math.cos(90-AV))
            sheet[f'J{i+2}'] = DH

            # * DV
            if (AV >= 90):
              DV = DH * math.tan(AV)
            else:
              DV = DH * math.tan(90-AV)
            sheet[f'K{i+2}'] = DV

            # * Cota B
            if (AV >= 90):
              CB = rowsExcel[i][0]+rowsExcel[i][1]-DV-rowsExcel[i][7]
            else:
              CB = rowsExcel[i][0]+rowsExcel[i][1]+DV-rowsExcel[i][7]
            sheet[f'N{i+2}'] = CB


          wb.save('NVGenerado.xlsx')
          print('Datos generados en NVGenerado.xlsx')


        procesoTerminado = True
        break
      else:
        print('-- Contraseña Incorrecta --')
    
  if (not userExist):
    print(f'-- El Usuario "{userInput}" No Existe --')
